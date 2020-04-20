import requests
import json
from bs4 import BeautifulSoup
from html.parser import HTMLParser
from openpyxl import Workbook
import time
import os

wb = Workbook()
ws =  wb.active
ws.title = "Data"

ws.cell(row=1, column=1, value="Time")
ws.cell(row=1, column=2, value="Price")
ws.cell(row=1, column=3, value="AllSummary")

ws.cell(row=1, column=5, value="RSI(14)")
ws.cell(row=1, column=6, value="STOCH(9,6)")
ws.cell(row=1, column=7, value="STOCHRSI(14)")
ws.cell(row=1, column=8, value="MACD(12,26)")
ws.cell(row=1, column=9, value="ADX(14)")
ws.cell(row=1, column=10, value="Williams %R")
ws.cell(row=1, column=11, value="CCI(14)")
ws.cell(row=1, column=12, value="ATR(14)")
ws.cell(row=1, column=13, value="Highs/Lows(14)")
ws.cell(row=1, column=14, value="Ultimate Oscillator")
ws.cell(row=1, column=15, value="ROC")
ws.cell(row=1, column=16, value="Bull/Bear Power(13)")
ws.cell(row=1, column=17, value="Buy")
ws.cell(row=1, column=18, value="Sell")
ws.cell(row=1, column=19, value="Neutral")
ws.cell(row=1, column=20, value="Techinal Summary")

ws.cell(row=1, column=22, value="5simple")
ws.cell(row=1, column=23, value="5expo")
ws.cell(row=1, column=24, value="10simple")
ws.cell(row=1, column=25, value="10expo")
ws.cell(row=1, column=26, value="20simple")
ws.cell(row=1, column=27, value="20expo")
ws.cell(row=1, column=28, value="50simple")
ws.cell(row=1, column=29, value="50expo")
ws.cell(row=1, column=30, value="100simple")
ws.cell(row=1, column=31, value="100expo")
ws.cell(row=1, column=32, value="200simple")
ws.cell(row=1, column=33, value="200expo")
ws.cell(row=1, column=34, value="Buy")
ws.cell(row=1, column=35, value="Sell")
ws.cell(row=1, column=36, value="Summary")

ws.cell(row=1, column=38, value='Clas S1')
ws.cell(row=1, column=39, value='Fib S1')
ws.cell(row=1, column=40, value='Cam S1')
ws.cell(row=1, column=41, value='Woo S1')
ws.cell(row=1, column=42, value='DeM s1')

ws.cell(row=1, column=44, value='Pivot')
ws.cell(row=1, column=45, value='Pivot')
ws.cell(row=1, column=46, value='Pivot')
ws.cell(row=1, column=47, value='Pivot')
ws.cell(row=1, column=48, value='Pivot')

ws.cell(row=1, column=50, value='priceNow')

ws.cell(row=1, column=52, value='clas R1')
ws.cell(row=1, column=53, value='Fib R1')
ws.cell(row=1, column=54, value='Cam R1')
ws.cell(row=1, column=55, value='Woo R1')
ws.cell(row=1, column=56, value='DeM R1')

numberOfIteration = 4000
sleepTime = 3
currencieType = 8839

for count in range(numberOfIteration):
    url = 'https://www.investing.com/indices/us-spx-500-futures-technical'
    payload = {
    "tab" : "indices",
    "options[periods][0]": 60,
    "options[periods][1]": 300,
    "options[periods][2]": 900,
    "options[periods][3]": 3600,
    "options[email_hour]": 17,
    "options[timezone_offset]": 19800,
    "options[currencies][]": currencieType
    }
    # Adding empty header as parameters are being sent in payload
    headers = {
    "Accept":"application/json, text/javascript, */*; q=0.01",
    "Accept-Encoding":"gzip, deflate, br",
    "Accept-Language":"en-IN,en;q=0.9,kn-IN;q=0.8,kn;q=0.7,en-US;q=0.6",
    "Connection":"keep-alive",
    "Content-Length":'202',
    "Content-Type":"application/x-www-form-urlencoded",
    "Host":"www.investing.com",
    "Origin":"https://www.investing.com",
    "Referer":"https://www.investing.com/indices/us-spx-500-futures-technical",
    "Sec-Fetch-Dest":"document",
    "Sec-Fetch-Mode":"navigate",
    "Sec-Fetch-Site":"same-origin",
    "User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.163 Safari/537.36",
    "X-Requested-With":"XMLHttpRequest"
    }

    r = requests.get(url, data=payload, headers=headers)
    # print(r.content)

    # print("JUST R> CONTENT ENDS HERRE ..............................")
    # print("Print each key-value pair from JSON response")
    # for key, value in r.json().items():
    #     print(key, ":", value)
    # #print(r.content)
    # print("<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<")

    pivotPoints = []
    technicalIndicatiors = []
    movingAverage = []
    allTogetherSummary = []
    stockPullTime = time.strftime("%H:%M:%S")

    soup = BeautifulSoup(r.content, 'html.parser')
    stockCurrTickPrice = float(soup.find(class_='arial_26 inlineblock pid-8839-last').text.replace(',', ''))
    pivotPointsData = soup.find(class_='genTbl closedTbl crossRatesTbl')
    technicalIndicatiorsData = soup.find(class_='halfSizeColumn float_lang_base_1')
    movingAverageData = soup.find(class_='halfSizeColumn float_lang_base_2')
    allTogetherSummaryData = soup.find(class_='summary').text
    allTogetherSummaryData = allTogetherSummaryData[8:]
    
    for string in pivotPointsData.stripped_strings:
        pivotPoints.append(string)

    for string in technicalIndicatiorsData.stripped_strings:
        technicalIndicatiors.append(string)

    for string in movingAverageData.stripped_strings:
        movingAverage.append(string)

    ws.cell(row=count+2, column=1, value=stockPullTime)
    ws.cell(row=count+2, column=2, value=stockCurrTickPrice)
    ws.cell(row=count+2, column=3, value=allTogetherSummaryData)

    ws.cell(row=count+2, column=5, value=technicalIndicatiors[7])
    ws.cell(row=count+2, column=6, value=technicalIndicatiors[10])
    ws.cell(row=count+2, column=7, value=technicalIndicatiors[13])
    ws.cell(row=count+2, column=8, value=technicalIndicatiors[16])
    ws.cell(row=count+2, column=9, value=technicalIndicatiors[19])
    ws.cell(row=count+2, column=10, value=technicalIndicatiors[22])
    ws.cell(row=count+2, column=11, value=technicalIndicatiors[25])
    ws.cell(row=count+2, column=12, value=technicalIndicatiors[28])
    ws.cell(row=count+2, column=13, value=technicalIndicatiors[31])
    ws.cell(row=count+2, column=14, value=technicalIndicatiors[34])
    ws.cell(row=count+2, column=15, value=technicalIndicatiors[37])
    ws.cell(row=count+2, column=16, value=technicalIndicatiors[40])
    ws.cell(row=count+2, column=17, value=int(technicalIndicatiors[42]))
    ws.cell(row=count+2, column=18, value=int(technicalIndicatiors[44]))
    ws.cell(row=count+2, column=19, value=int(technicalIndicatiors[46]))
    ws.cell(row=count+2, column=20, value=technicalIndicatiors[48])

    ws.cell(row=count+2, column=22, value=movingAverage[7])
    ws.cell(row=count+2, column=23, value=movingAverage[9])
    ws.cell(row=count+2, column=24, value=movingAverage[12])
    ws.cell(row=count+2, column=25, value=movingAverage[14])
    ws.cell(row=count+2, column=26, value=movingAverage[17])
    ws.cell(row=count+2, column=27, value=movingAverage[19])
    ws.cell(row=count+2, column=28, value=movingAverage[22])
    ws.cell(row=count+2, column=29, value=movingAverage[24])
    ws.cell(row=count+2, column=30, value=movingAverage[27])
    ws.cell(row=count+2, column=31, value=movingAverage[29])
    ws.cell(row=count+2, column=32, value=movingAverage[32])
    ws.cell(row=count+2, column=33, value=movingAverage[34])
    ws.cell(row=count+2, column=34, value=int(movingAverage[36]))
    ws.cell(row=count+2, column=35, value=int(movingAverage[38]))
    ws.cell(row=count+2, column=36, value=movingAverage[40])

    ws.cell(row=count+2, column=38, value=float(pivotPoints[11].replace(',', '')))
    ws.cell(row=count+2, column=39, value=float(pivotPoints[19].replace(',', '')))
    ws.cell(row=count+2, column=40, value=float(pivotPoints[27].replace(',', '')))
    ws.cell(row=count+2, column=41, value=float(pivotPoints[35].replace(',', '')))
    ws.cell(row=count+2, column=42, value=float(pivotPoints[43].replace(',', '')))

    ws.cell(row=count+2, column=44, value=float(pivotPoints[12].replace(',', '')))
    ws.cell(row=count+2, column=45, value=float(pivotPoints[20].replace(',', '')))
    ws.cell(row=count+2, column=46, value=float(pivotPoints[28].replace(',', '')))
    ws.cell(row=count+2, column=47, value=float(pivotPoints[36].replace(',', '')))
    ws.cell(row=count+2, column=48, value=float(pivotPoints[44].replace(',', '')))

    ws.cell(row=count+2, column=50, value=stockCurrTickPrice)

    ws.cell(row=count+2, column=52, value=float(pivotPoints[13].replace(',', '')))
    ws.cell(row=count+2, column=53, value=float(pivotPoints[21].replace(',', '')))
    ws.cell(row=count+2, column=54, value=float(pivotPoints[29].replace(',', '')))
    ws.cell(row=count+2, column=55, value=float(pivotPoints[37].replace(',', '')))
    ws.cell(row=count+2, column=56, value=float(pivotPoints[45].replace(',', '')))

    print(count,'. ', stockPullTime, ':', stockCurrTickPrice, ':', allTogetherSummaryData)

    cwd = os.path.split(os.path.abspath(__file__))
    excelFileName = cwd[1][:-3] + ".xlsx"
    wb.save(filename = cwd[0] + '/' + excelFileName)
    
    time.sleep(sleepTime)

print("Hola! Done B! data has been put to excel")
