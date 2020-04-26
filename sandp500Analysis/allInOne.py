import requests
import json
from bs4 import BeautifulSoup
from html.parser import HTMLParser
from openpyxl import Workbook
import time
import os
import openpyxl
import queue

cwd = os.path.split(os.path.abspath(__file__))
excelFileName = cwd[1][:-3] + ".xlsx"

path = cwd[0] + '\\'+ excelFileName
wb = openpyxl.load_workbook(path)

savePath = cwd[0] + '\\'+ 'checkOut.xlsx'
anotherfile = openpyxl.load_workbook(savePath)

tenHourlySummery = 0
counterToStore = 0

ws = wb.active

maxColumn = ws.max_row
maxColumn = maxColumn + 6

ws.cell(row=maxColumn-2, column=1, value="NewData")
ws.cell(row=maxColumn-2, column=3, value="From")
ws.cell(row=maxColumn-2, column=4, value=time.strftime("%H:%M"))
ws.cell(row=maxColumn-2, column=6, value="To")
ws.cell(row=maxColumn-2, column=8, value="On")
named_tuple = time.localtime()
ws.cell(row=maxColumn-2, column=9, value=time.strftime("%d %b %Y", named_tuple))

ws.cell(row=1, column=1, value="Time")
ws.cell(row=1, column=2, value="Price")
ws.cell(row=1, column=3, value="AllSummary")
#
ws.cell(row=1, column=5, value="1min MA")
ws.cell(row=1, column=6, value="1min Ind")
ws.cell(row=1, column=7, value="1min Sum")

ws.cell(row=1, column=9, value="5min MA")
ws.cell(row=1, column=10, value="5min Ind")
ws.cell(row=1, column=11, value="5min Sum")

ws.cell(row=1, column=13, value="15min MA")
ws.cell(row=1, column=14, value="15min Ind")
ws.cell(row=1, column=15, value="15min Sum")

ws.cell(row=1, column=17, value="1hour MA")
ws.cell(row=1, column=18, value="1hour Ind")
ws.cell(row=1, column=19, value="1hour Sum")
#
ws.cell(row=1, column=21, value="RSI(14)")
ws.cell(row=1, column=22, value="STOCH(9,6)")
ws.cell(row=1, column=23, value="STOCHRSI(14)")
ws.cell(row=1, column=24, value="MACD(12,26)")
ws.cell(row=1, column=25, value="ADX(14)")
ws.cell(row=1, column=26, value="Williams %R")
ws.cell(row=1, column=27, value="CCI(14)")
ws.cell(row=1, column=28, value="ATR(14)")
ws.cell(row=1, column=29, value="Highs/Lows(14)")
ws.cell(row=1, column=30, value="Ultimate Oscillator")
ws.cell(row=1, column=31, value="ROC")
ws.cell(row=1, column=32, value="Bull/Bear Power(13)")
ws.cell(row=1, column=33, value="Buy")
ws.cell(row=1, column=34, value="Sell")
ws.cell(row=1, column=35, value="Neutral")
ws.cell(row=1, column=36, value="Techinal Summary")

ws.cell(row=1, column=38, value="5simple")
ws.cell(row=1, column=39, value="5expo")
ws.cell(row=1, column=40, value="10simple")
ws.cell(row=1, column=41, value="10expo")
ws.cell(row=1, column=42, value="20simple")
ws.cell(row=1, column=43, value="20expo")
ws.cell(row=1, column=44, value="50simple")
ws.cell(row=1, column=45, value="50expo")
ws.cell(row=1, column=46, value="100simple")
ws.cell(row=1, column=47, value="100expo")
ws.cell(row=1, column=48, value="200simple")
ws.cell(row=1, column=49, value="200expo")
ws.cell(row=1, column=50, value="Buy")
ws.cell(row=1, column=51, value="Sell")
ws.cell(row=1, column=52, value="MA Summary")

ws.cell(row=1, column=54, value='Clas S1')
ws.cell(row=1, column=55, value='Fib S1')
ws.cell(row=1, column=56, value='Cam S1')
ws.cell(row=1, column=57, value='Woo S1')
ws.cell(row=1, column=58, value='DeM s1')

ws.cell(row=1, column=60, value='Pivot')
ws.cell(row=1, column=61, value='Pivot')
ws.cell(row=1, column=62, value='Pivot')
ws.cell(row=1, column=63, value='Pivot')
ws.cell(row=1, column=64, value='Pivot')

ws.cell(row=1, column=66, value='priceNow')

ws.cell(row=1, column=68, value='clas R1')
ws.cell(row=1, column=69, value='Fib R1')
ws.cell(row=1, column=70, value='Cam R1')
ws.cell(row=1, column=71, value='Woo R1')
ws.cell(row=1, column=72, value='DeM R1')

numberOfIteration = 20000
sleepTime = 3
currencieType = 8839

urlAll = 'https://www.investing.com/technical/Service/GetSummaryTable'
payloadAll = {
    "tab" : "indices",
    "options[periods][0]": 60,
    "options[periods][1]": 300,
    "options[periods][2]": 900,
    "options[periods][3]": 3600,
    "options[email_hour]": 17,
    "options[timezone_offset]": 19800,
    "options[currencies][]": currencieType
}
# Adding empty header as parameters are being sent in payload
headersAll = {
    "Accept":"application/json, text/javascript, */*; q=0.01",
    "Accept-Encoding":"gzip, deflate, br",
    "Accept-Language":"en-IN,en;q=0.9,kn-IN;q=0.8,kn;q=0.7,en-US;q=0.6",
    "Connection":"keep-alive",
    "Content-Length":202,
    "Content-Type":"application/x-www-form-urlencoded",
    "Host":"www.investing.com",
    "Origin":"https://www.investing.com",
    "Referer":"https://www.investing.com/technical/technical-summary",
    "Sec-Fetch-Dest":"empty",
    "Sec-Fetch-Mode":"cors",
    "Sec-Fetch-Site":"same-origin",
    "User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.163 Safari/537.36",
    "X-Requested-With":"XMLHttpRequest"
}

url = 'https://www.investing.com/indices/us-spx-500-futures-technical'
payload = {
"tab" : "indices",
"options[periods][0]": 60,
"options[periods][1]": 300,
"options[periods][2]": 900,
"options[periods][3]": 3600,
"options[email_hour]": 17,
"options[timezone_offset]": 19800,
"options[currencies][]": currencieType
}
# Adding empty header as parameters are being sent in payload
headers = {
"Accept":"application/json, text/javascript, */*; q=0.01",
"Accept-Encoding":"gzip, deflate, br",
"Accept-Language":"en-IN,en;q=0.9,kn-IN;q=0.8,kn;q=0.7,en-US;q=0.6",
"Connection":"keep-alive",
"Content-Length":'202',
"Content-Type":"application/x-www-form-urlencoded",
"Host":"www.investing.com",
"Origin":"https://www.investing.com",
"Referer":"https://www.investing.com/indices/us-spx-500-futures-technical",
"Sec-Fetch-Dest":"document",
"Sec-Fetch-Mode":"navigate",
"Sec-Fetch-Site":"same-origin",
"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.163 Safari/537.36",
"X-Requested-With":"XMLHttpRequest"
}

for count in range(maxColumn, numberOfIteration):
    r = requests.get(url, data=payload, headers=headers)
    # print(r.content)

    # print("JUST R> CONTENT ENDS HERRE ..............................")
    # print("Print each key-value pair from JSON response")
    # for key, value in r.json().items():
    #     print(key, ":", value)
    # #print(r.content)
    # print("<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<")

    pivotPoints = []
    technicalIndicatiors = []
    movingAverage = []
    allTogetherSummary = []
    stockPullTime = time.strftime("%H:%M:%S")

    soup = BeautifulSoup(r.content, 'html.parser')
    stockCurrTickPrice = float(soup.find(class_='arial_26 inlineblock pid-8839-last').text.replace(',', ''))
    pivotPointsData = soup.find(class_='genTbl closedTbl crossRatesTbl')
    technicalIndicatiorsData = soup.find(class_='halfSizeColumn float_lang_base_1')
    movingAverageData = soup.find(class_='halfSizeColumn float_lang_base_2')
    allTogetherSummaryData = soup.find(class_='summary').text
    allTogetherSummaryData = allTogetherSummaryData[8:]
    
    for string in pivotPointsData.stripped_strings:
        pivotPoints.append(string)

    for string in technicalIndicatiorsData.stripped_strings:
        technicalIndicatiors.append(string)

    for string in movingAverageData.stripped_strings:
        movingAverage.append(string)

    #
    rAll = requests.post(urlAll, data=payloadAll, headers=headersAll)

    tableHead =[]
    tableBody =[]

    soup = BeautifulSoup(rAll.json()["html"], "html.parser")

    for string in soup.thead.stripped_strings:
        tableHead.append(string)

    for string in soup.tbody.stripped_strings:
        tableBody.append(string)

    stockName = tableBody[0]
    stockCurrTickPrice = tableBody[1]
    stockPullTime = time.strftime("%H:%M:%S")

    tableHead = tableHead [2:]
    tableBody = tableBody [2:]
    #

    ws.cell(row=count, column=1, value=stockPullTime)
    ws.cell(row=count, column=2, value=float(stockCurrTickPrice.replace(',', '')))
    ws.cell(row=count, column=3, value=allTogetherSummaryData)

    ws.cell(row=count, column=5, value=tableBody[1])
    ws.cell(row=count, column=6, value=tableBody[6])
    ws.cell(row=count, column=7, value=tableBody[11])

    ws.cell(row=count, column=9, value=tableBody[2])
    ws.cell(row=count, column=10, value=tableBody[7])
    ws.cell(row=count, column=11, value=tableBody[12])

    ws.cell(row=count, column=13, value=tableBody[3])
    ws.cell(row=count, column=14, value=tableBody[8])
    ws.cell(row=count, column=15, value=tableBody[13])

    ws.cell(row=count, column=17, value=tableBody[4])
    ws.cell(row=count, column=18, value=tableBody[9])
    ws.cell(row=count, column=19, value=tableBody[14])

    ws.cell(row=count, column=21, value=technicalIndicatiors[7])
    ws.cell(row=count, column=22, value=technicalIndicatiors[10])
    ws.cell(row=count, column=23, value=technicalIndicatiors[13])
    ws.cell(row=count, column=24, value=technicalIndicatiors[16])
    ws.cell(row=count, column=25, value=technicalIndicatiors[19])
    ws.cell(row=count, column=26, value=technicalIndicatiors[22])
    ws.cell(row=count, column=27, value=technicalIndicatiors[25])
    ws.cell(row=count, column=28, value=technicalIndicatiors[28])
    ws.cell(row=count, column=29, value=technicalIndicatiors[31])
    ws.cell(row=count, column=30, value=technicalIndicatiors[34])
    ws.cell(row=count, column=31, value=technicalIndicatiors[37])
    ws.cell(row=count, column=32, value=technicalIndicatiors[40])
    ws.cell(row=count, column=33, value=int(technicalIndicatiors[42]))
    ws.cell(row=count, column=34, value=int(technicalIndicatiors[44]))
    ws.cell(row=count, column=35, value=int(technicalIndicatiors[46]))
    ws.cell(row=count, column=36, value=technicalIndicatiors[48])

    ws.cell(row=count, column=38, value=movingAverage[7])
    ws.cell(row=count, column=39, value=movingAverage[9])
    ws.cell(row=count, column=40, value=movingAverage[12])
    ws.cell(row=count, column=41, value=movingAverage[14])
    ws.cell(row=count, column=42, value=movingAverage[17])
    ws.cell(row=count, column=43, value=movingAverage[19])
    ws.cell(row=count, column=44, value=movingAverage[22])
    ws.cell(row=count, column=45, value=movingAverage[24])
    ws.cell(row=count, column=46, value=movingAverage[27])
    ws.cell(row=count, column=47, value=movingAverage[29])
    ws.cell(row=count, column=48, value=movingAverage[32])
    ws.cell(row=count, column=49, value=movingAverage[34])
    ws.cell(row=count, column=50, value=int(movingAverage[36]))
    ws.cell(row=count, column=51, value=int(movingAverage[38]))
    ws.cell(row=count, column=52, value=movingAverage[40])

    ws.cell(row=count, column=54, value=float(pivotPoints[11].replace(',', '')))
    ws.cell(row=count, column=55, value=float(pivotPoints[19].replace(',', '')))
    ws.cell(row=count, column=56, value=float(pivotPoints[27].replace(',', '')))
    ws.cell(row=count, column=57, value=float(pivotPoints[35].replace(',', '')))
    ws.cell(row=count, column=58, value=float(pivotPoints[43].replace(',', '')))

    ws.cell(row=count, column=60, value=float(pivotPoints[12].replace(',', '')))
    ws.cell(row=count, column=61, value=float(pivotPoints[20].replace(',', '')))
    ws.cell(row=count, column=62, value=float(pivotPoints[28].replace(',', '')))
    ws.cell(row=count, column=63, value=float(pivotPoints[36].replace(',', '')))
    ws.cell(row=count, column=64, value=float(pivotPoints[44].replace(',', '')))

    ws.cell(row=count, column=66, value=stockCurrTickPrice)

    ws.cell(row=count, column=68, value=float(pivotPoints[13].replace(',', '')))
    ws.cell(row=count, column=69, value=float(pivotPoints[21].replace(',', '')))
    ws.cell(row=count, column=70, value=float(pivotPoints[29].replace(',', '')))
    ws.cell(row=count, column=71, value=float(pivotPoints[37].replace(',', '')))
    ws.cell(row=count, column=72, value=float(pivotPoints[45].replace(',', '')))

    #tenHourlySummery begins here

    if (tableBody[14] == 'Neutral'):
        tenHourlySummery = tenHourlySummery + 1
    else:
        tenHourlySummery = 0
    
    if (tenHourlySummery >= 10):
        if(tableBody[14] != "Neutral"):
            counterToStore = counterToStore + 1
            newFile = anotherfile.active
            newFile.cell(row=counterToStore, column=1, value=stockPullTime)
            newFile.cell(row=counterToStore, column=3, value=tableBody[14])
            newFile.cell(row=counterToStore, column=2, value=stockCurrTickPrice)
            cwd = os.path.split(os.path.abspath(__file__))
            anotherfile.save(filename = cwd[0] + '/' + 'checkOut.xlsx')
            tenHourlySummery = 0
            ws = wb.active
            print("Entry Done !")

    #tenHourlySummery ends here
    ws.cell(row=maxColumn-2, column=7, value=time.strftime("%H:%M"))
    print(counterToStore, ". ", stockPullTime, "-", "-", stockCurrTickPrice, "\t1ms=", tableBody[11], "\t5ms=", tableBody[12], "\t15ms=", tableBody[13], "\t1Hs=", tableBody[14], "\tTechnicalSumm=", technicalIndicatiors[48], "\tMovingSummary=",movingAverage[40])
    
    cwd = os.path.split(os.path.abspath(__file__))
    excelFileName = cwd[1][:-3] + ".xlsx"
    wb.save(filename = cwd[0] + '/' + excelFileName)
    
    time.sleep(sleepTime)

print("Hola! Done B! data has been put to excel")
