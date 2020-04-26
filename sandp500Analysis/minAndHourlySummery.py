import requests
import json
from bs4 import BeautifulSoup
from html.parser import HTMLParser
from openpyxl import Workbook
import time
import os

wb = Workbook()
ws =  wb.active
ws.title = "Data"

ws.cell(row=1, column=1, value="Time")
ws.cell(row=1, column=2, value="Price")

ws.cell(row=1, column=4, value="1min MA")
ws.cell(row=1, column=5, value="1min Ind")
ws.cell(row=1, column=6, value="1min Sum")

ws.cell(row=1, column=8, value="5min MA")
ws.cell(row=1, column=9, value="5min Ind")
ws.cell(row=1, column=10, value="5min Sum")

ws.cell(row=1, column=12, value="15min MA")
ws.cell(row=1, column=13, value="15min Ind")
ws.cell(row=1, column=14, value="15min Sum")

ws.cell(row=1, column=16, value="1hour MA")
ws.cell(row=1, column=17, value="1hour Ind")
ws.cell(row=1, column=18, value="1hour Sum")

numberOfIteration = 4000
sleepTime = 4
currencieType = 8839

for count in range(numberOfIteration):
    url = 'https://www.investing.com/technical/Service/GetSummaryTable'
    payload = {
        "tab" : "indices",
        "options[periods][0]": 60,
        "options[periods][1]": 300,
        "options[periods][2]": 900,
        "options[periods][3]": 3600,
        "options[email_hour]": 17,
        "options[timezone_offset]": 19800,
        "options[currencies][]": currencieType
    }
    # Adding empty header as parameters are being sent in payload
    headers = {
        "Accept":"application/json, text/javascript, */*; q=0.01",
        "Accept-Encoding":"gzip, deflate, br",
        "Accept-Language":"en-IN,en;q=0.9,kn-IN;q=0.8,kn;q=0.7,en-US;q=0.6",
        "Connection":"keep-alive",
        "Content-Length":202,
        "Content-Type":"application/x-www-form-urlencoded",
        "Host":"www.investing.com",
        "Origin":"https://www.investing.com",
        "Referer":"https://www.investing.com/technical/technical-summary",
        "Sec-Fetch-Dest":"empty",
        "Sec-Fetch-Mode":"cors",
        "Sec-Fetch-Site":"same-origin",
        "User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.163 Safari/537.36",
        "X-Requested-With":"XMLHttpRequest"
    }

    r = requests.post(url, data=payload, headers=headers)
    # print(r.content)

    # print("JUST R> CONTENT ENDS HERRE ..............................")
    # print("Print each key-value pair from JSON response")
    # for key, value in r.json().items():
    #     print(key, ":", value)
    # #print(r.content)
    # print("<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<")

    #print (r.json()["html"])

    tableHead =[]
    tableBody =[]

    soup = BeautifulSoup(r.json()["html"], "html.parser")

    for string in soup.thead.stripped_strings:
        tableHead.append(string)

    for string in soup.tbody.stripped_strings:
        tableBody.append(string)

    stockName = tableBody[0]
    stockCurrTickPrice = tableBody[1]
    stockPullTime = time.strftime("%H:%M:%S")

    tableHead = tableHead [2:]
    tableBody = tableBody [2:]

    # # Creates Json if needed
    # obj = [{
    #     tableHead[0]:{
    #     tableBody[0] : tableBody[1],
    #     tableBody[5] : tableBody[6],
    #     tableBody[10]  : tableBody[11]
    # }},
    # {  
    #     tableHead[1]:{
    #     tableBody[0] : tableBody[2],
    #     tableBody[5] : tableBody[7],
    #     tableBody[10]  : tableBody[12]
    # }},
    # {  
    #     tableHead[2]:{
    #     tableBody[0] : tableBody[3],
    #     tableBody[5] : tableBody[8],
    #     tableBody[10]  : tableBody[13]
    # }},
    # {  
    #     tableHead[3]:{
    #     tableBody[0] : tableBody[4],
    #     tableBody[5] : tableBody[9],
    #     tableBody[10]  : tableBody[14]
    # }}
    # ]

    #####################################################
    # Data in obj 
    # [
    #   {
    #     "1 Minute": {
    #       "Moving Averages:": "Strong Buy",
    #       "Indicators:": "Strong Buy",
    #       "Summary:": "Strong Buy"
    #     }
    #   },
    #   {
    #     "5 Minutes": {
    #       "Moving Averages:": "Strong Buy",
    #       "Indicators:": "Strong Buy",
    #       "Summary:": "Strong Buy"
    #     }
    #   },
    #   {
    #     "15 Minutes": {
    #       "Moving Averages:": "Strong Buy",
    #       "Indicators:": "Strong Buy",
    #       "Summary:": "Strong Buy"
    #     }
    #   },
    #   {
    #     "Hourly": {
    #       "Moving Averages:": "Strong Buy",
    #       "Indicators:": "Strong Buy",
    #       "Summary:": "Strong Buy"
    #     }
    #   }
    # ]
    #
    # Possible Values Include: 
    # Strong Sell, Sell
    # Neutral
    # Strong Buy, Buy
    #####################################################

    # print(obj[3]['Daily']['Indicators:'])
    # print(stockName, " : ", stockCurrTickPrice)
    # print(json.dumps(obj, indent=2))
    
    ws.cell(row=count+2, column=1, value=stockPullTime)
    ws.cell(row=count+2, column=2, value=float(stockCurrTickPrice.replace(',', '')))

    ws.cell(row=count+2, column=4, value=tableBody[1])
    ws.cell(row=count+2, column=5, value=tableBody[6])
    ws.cell(row=count+2, column=6, value=tableBody[11])

    ws.cell(row=count+2, column=8, value=tableBody[2])
    ws.cell(row=count+2, column=9, value=tableBody[7])
    ws.cell(row=count+2, column=10, value=tableBody[12])

    ws.cell(row=count+2, column=12, value=tableBody[3])
    ws.cell(row=count+2, column=13, value=tableBody[8])
    ws.cell(row=count+2, column=14, value=tableBody[13])

    ws.cell(row=count+2, column=16, value=tableBody[4])
    ws.cell(row=count+2, column=17, value=tableBody[9])
    ws.cell(row=count+2, column=18, value=tableBody[14])

    # just for visual
    print(count+1, ". ", stockPullTime, "-", stockName, "-", stockCurrTickPrice, "\t1ms=", tableBody[11], "\t5ms=", tableBody[12], "\t15ms=", tableBody[13], "\t1Hs=", tableBody[14])
    
    cwd = os.path.split(os.path.abspath(__file__))
    excelFileName = cwd[1][:-3] + ".xlsx"
    wb.save(filename = cwd[0] + '/' + excelFileName)
    
    time.sleep(sleepTime)

print("Hola! Done! data has been put to excel")
