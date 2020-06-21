import requests
import bs4
from openpyxl import Workbook
import os
import time
from datetime import datetime
from pytz import timezone  

wb = Workbook()
ws1 =  wb.active
ws1.title = "Data"
ws2 = wb.create_sheet()
ws2.title = "Data2"
ittRange = 6
symbolValue = "bank-nifty"
replaceDict = {"Strong Sell":"S-Sell", "Sell":"Sell", "Neutral":"Neutral", "Buy":"Buy", "Strong Buy":"S-Buy",}

#Sheet 1
ws1.cell(row=2, column=1, value="Time")
ws1.cell(row=2, column=2, value="Price")

ws1.cell(row=1, column=4, value="1Min")
ws1.cell(row=2, column=4, value="Ma")
ws1.cell(row=2, column=5, value="Momt O")
ws1.cell(row=2, column=6, value="Trend O")
ws1.cell(row=2, column=7, value="Volatility")

ws1.cell(row=1, column=9, value="5Min")
ws1.cell(row=2, column=9, value="Ma")
ws1.cell(row=2, column=10, value="Momt O")
ws1.cell(row=2, column=11, value="Trend O")
ws1.cell(row=2, column=12, value="Volatility")

ws1.cell(row=1, column=14, value="15Min")
ws1.cell(row=2, column=14, value="Ma")
ws1.cell(row=2, column=15, value="Momt O")
ws1.cell(row=2, column=16, value="Trend O")
ws1.cell(row=2, column=17, value="Volatility")

ws1.cell(row=1, column=19, value="1H")
ws1.cell(row=2, column=19, value="Ma")
ws1.cell(row=2, column=20, value="Momt O")
ws1.cell(row=2, column=21, value="Trend O")
ws1.cell(row=2, column=22, value="Volatility")

ws1.cell(row=1, column=24, value="1D")
ws1.cell(row=2, column=24, value="Ma")
ws1.cell(row=2, column=25, value="Momt O")
ws1.cell(row=2, column=26, value="Trend O")
ws1.cell(row=2, column=27, value="Volatility")


#Sheet 2
ws2.cell(row=2, column=1, value="Time")
ws2.cell(row=2, column=2, value="Price")

ws2.cell(row=1, column=4, value="MA")
ws2.cell(row=2, column=4, value="1Min")
ws2.cell(row=2, column=5, value="5min")
ws2.cell(row=2, column=6, value="15Min")
ws2.cell(row=2, column=7, value="1H")
ws2.cell(row=2, column=8, value="1D")

ws2.cell(row=1, column=10, value="Momt O")
ws2.cell(row=2, column=10, value="1Min")
ws2.cell(row=2, column=11, value="5min")
ws2.cell(row=2, column=12, value="15Min")
ws2.cell(row=2, column=13, value="1H")
ws2.cell(row=2, column=14, value="1D")

ws2.cell(row=1, column=16, value="Trend O")
ws2.cell(row=2, column=16, value="1Min")
ws2.cell(row=2, column=17, value="5min")
ws2.cell(row=2, column=18, value="15Min")
ws2.cell(row=2, column=19, value="1H")
ws2.cell(row=2, column=20, value="1D")

ws2.cell(row=1, column=22, value="Volatility")
ws2.cell(row=2, column=22, value="1Min")
ws2.cell(row=2, column=23, value="5min")
ws2.cell(row=2, column=24, value="15Min")
ws2.cell(row=2, column=25, value="1H")
ws2.cell(row=2, column=26, value="1D")

for i in range(1,ittRange):

    oneMinData = []
    tempData = []
    timeFrame = ['60', '300', '900', '3600', '86400']
    headers = {
    "Accept":"text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9",
    "Accept-Encoding":"gzip, deflate, br",
    "Accept-Language":"en-IN,en;q=0.9,kn-IN;q=0.8,kn;q=0.7,en-US;q=0.6",
    "Cache-Control":"max-age=0",
    "Connection":"keep-alive",
    "Host":"in.investing.com",
    "Referer":"https://in.investing.com/indices/s-p-cnx-nifty-news",
    "Sec-Fetch-Dest":"document",
    "Sec-Fetch-Mode":"navigate",
    "Sec-Fetch-Site":"same-origin",
    "Sec-Fetch-User":"?1",
    "Upgrade-Insecure-Requests":"1",
    "User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.106 Safari/537.36"
    }

    for j in range(0,5):
        url = "https://in.investing.com/indices/" + symbolValue + "-technical?timeFrame=" + timeFrame[j]
        dataOne = requests.get(url,headers=headers)
        soup = bs4.BeautifulSoup(dataOne.text,"lxml")

        if(j == 0):
            pricePick = soup.select('.last-price-value.js-streamable-element')
            pricePick = pricePick[0].text

        soup = soup.select('.common-table-comp.e-technical-summary-table')
        soup = soup[0].table

        for row in soup.find_all('tr'):
            for cell in row.find_all('td'):
                tempData.append(cell.text)
        
        oneMinData.append(tempData)
        tempData = []

    rigntNow = datetime.now(timezone('asia/kolkata')).strftime('%H:%M:%S')

    #Sheet 1
    ws1.cell(row=i+2, column=1, value=rigntNow)
    ws1.cell(row=i+2, column=2, value=float(pricePick.replace(',', '')))

    ws1.cell(row=i+2, column=4, value=replaceDict[oneMinData[0][4]])
    ws1.cell(row=i+2, column=5, value=replaceDict[oneMinData[0][9]])
    ws1.cell(row=i+2, column=6, value=replaceDict[oneMinData[0][14]])
    ws1.cell(row=i+2, column=7, value=replaceDict[oneMinData[0][19]])

    ws1.cell(row=i+2, column=9, value=replaceDict[oneMinData[1][4]])
    ws1.cell(row=i+2, column=10, value=replaceDict[oneMinData[1][9]])
    ws1.cell(row=i+2, column=11, value=replaceDict[oneMinData[1][14]])
    ws1.cell(row=i+2, column=12, value=replaceDict[oneMinData[1][19]])

    ws1.cell(row=i+2, column=14, value=replaceDict[oneMinData[2][4]])
    ws1.cell(row=i+2, column=15, value=replaceDict[oneMinData[2][9]])
    ws1.cell(row=i+2, column=16, value=replaceDict[oneMinData[2][14]])
    ws1.cell(row=i+2, column=17, value=replaceDict[oneMinData[2][19]])

    ws1.cell(row=i+2, column=19, value=replaceDict[oneMinData[3][4]])
    ws1.cell(row=i+2, column=20, value=replaceDict[oneMinData[3][9]])
    ws1.cell(row=i+2, column=21, value=replaceDict[oneMinData[3][14]])
    ws1.cell(row=i+2, column=22, value=replaceDict[oneMinData[3][19]])

    ws1.cell(row=i+2, column=24, value=replaceDict[oneMinData[4][4]])
    ws1.cell(row=i+2, column=25, value=replaceDict[oneMinData[4][9]])
    ws1.cell(row=i+2, column=26, value=replaceDict[oneMinData[4][14]])
    ws1.cell(row=i+2, column=27, value=replaceDict[oneMinData[4][19]])

    #Sheet 2
    ws2.cell(row=i+2, column=1, value=rigntNow)
    ws2.cell(row=i+2, column=2, value=float(pricePick.replace(',', '')))

    ws2.cell(row=i+2, column=4, value=replaceDict[oneMinData[0][4]])
    ws2.cell(row=i+2, column=5, value=replaceDict[oneMinData[1][4]])
    ws2.cell(row=i+2, column=6, value=replaceDict[oneMinData[2][4]])
    ws2.cell(row=i+2, column=7, value=replaceDict[oneMinData[3][4]])
    ws2.cell(row=i+2, column=8, value=replaceDict[oneMinData[4][4]])

    ws2.cell(row=i+2, column=10, value=replaceDict[oneMinData[0][9]])
    ws2.cell(row=i+2, column=11, value=replaceDict[oneMinData[1][9]])
    ws2.cell(row=i+2, column=12, value=replaceDict[oneMinData[2][9]])
    ws2.cell(row=i+2, column=13, value=replaceDict[oneMinData[3][9]])
    ws2.cell(row=i+2, column=14, value=replaceDict[oneMinData[4][9]])

    ws2.cell(row=i+2, column=16, value=replaceDict[oneMinData[0][14]])
    ws2.cell(row=i+2, column=17, value=replaceDict[oneMinData[1][14]])
    ws2.cell(row=i+2, column=18, value=replaceDict[oneMinData[2][14]])
    ws2.cell(row=i+2, column=19, value=replaceDict[oneMinData[3][14]])
    ws2.cell(row=i+2, column=20, value=replaceDict[oneMinData[4][19]])

    ws2.cell(row=i+2, column=22, value=replaceDict[oneMinData[0][19]])
    ws2.cell(row=i+2, column=23, value=replaceDict[oneMinData[1][19]])
    ws2.cell(row=i+2, column=24, value=replaceDict[oneMinData[2][19]])
    ws2.cell(row=i+2, column=25, value=replaceDict[oneMinData[3][19]])
    ws2.cell(row=i+2, column=26, value=replaceDict[oneMinData[4][14]])


    print(i, rigntNow, pricePick, "\t", replaceDict[oneMinData[0][4]], replaceDict[oneMinData[0][9]], replaceDict[oneMinData[0][14]], replaceDict[oneMinData[0][19]],
          "\t", replaceDict[oneMinData[1][4]], replaceDict[oneMinData[1][9]], replaceDict[oneMinData[1][14]], replaceDict[oneMinData[1][19]],
          "\t", replaceDict[oneMinData[2][4]], replaceDict[oneMinData[2][9]], replaceDict[oneMinData[2][14]], replaceDict[oneMinData[2][19]],
          "\t", replaceDict[oneMinData[3][4]], replaceDict[oneMinData[3][9]], replaceDict[oneMinData[3][14]], replaceDict[oneMinData[3][19]],
          "\t", replaceDict[oneMinData[4][4]], replaceDict[oneMinData[4][9]], replaceDict[oneMinData[4][14]], replaceDict[oneMinData[4][19]])
    cwd = os.path.split(os.path.abspath(__file__))
    excelFileName = cwd[1][:-3] + ".xlsx"
    wb.save(filename = cwd[0] + '/' + excelFileName)
    time.sleep(2)
    